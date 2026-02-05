import pypdf
import msoffcrypto
import io
import os
import secrets
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def protect_pdf(input_path, password, user_password="1234"):
    """
    Cifra un archivo PDF existente mediante una contraseña.
    
    Args:
        input_path (str): Ruta local del archivo PDF a proteger.
        password (str): Contraseña que se aplicará al archivo.
        user_password (str) dafault="1234": contraseña de usuario, aplica diferentes restricciones.
    """
    try:
        if not os.path.exists(input_path):
            print(f"⚠️ Error: No se encontró el PDF en {input_path}")
            return

        reader = pypdf.PdfReader(input_path)
        writer = pypdf.PdfWriter()

        for page in reader.pages:
            writer.add_page(page)

        writer.encrypt(user_password=user_password, owner_password=password, permissions_flag=-3904, algorithm="AES-256-R5")

        with open(input_path, "wb") as f:
            writer.write(f)
        
        print(f"🔒 PDF encriptado correctamente: {os.path.basename(input_path)}")
        
    except Exception as e:
        print(f"❌ Error al encriptar el PDF: {e}")

# Función de ayuda para proteger las hojas del excel. 
def protect_sheet_only(file_path, password="", blind=False):
    """
    Bloquea las celdas de un archivo xls. Estiliza el encabezado y
    autoajusta las columnas antes de aplicar el bloqueo de celdas
    Args:
        input_path (str): Ruta local del archivo PDF a proteger.
        password (str): Contraseña que se aplicará al archivo.
        blind (Bool)default=False: Crea una contraseña ciega de cifrado
                que nadie conoce mediante el modulo secrets. 
    """
    wb = load_workbook(file_path)
    target_password = secrets.token_hex(8) if blind else password
    # Definir el estilo del encabezado (Azul profesional)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center")

    for ws in wb.worksheets:
        # 1. ESTILO DE ENCABEZADOS (Fila 1)
        for cell in ws[1]: # Itera sobre la primera fila
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 2. AUTOAJUSTE DE COLUMNAS
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter 
            
            for cell in column:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            
            # Ajuste dinámico: ancho del texto + un pequeño margen
            ws.column_dimensions[column_letter].width = max_length + 3

        # 3. BLOQUEO FINAL
        # Permitimos 'formatColumns' para que el usuario pueda estirar columnas si quiere
        ws.protection.set_password(target_password)
        ws.protection.sheet = True
        ws.protection.formatColumns = False # Cambiar a True si quieres que el usuario las pueda estirar
        
    wb.save(file_path)
# Función de ayuda para encriptar en 2 niveles.
def encryp_excel(file_path, password):
    """
    Cifra un archivo Excel (.xlsx) mediante el estándar de Microsoft Office.
    Args:
        file_path (str): Ruta local del archivo Excel a proteger.
        password (str): Contraseña que se aplicará al archivo.
    """
    try:
        if not os.path.exists(file_path):
            print(f"⚠️ Error: No se encontró el Excel en {file_path}")
            return

        # Leer el archivo original en binario
        with open(file_path, "rb") as f:
            file_data = io.BytesIO(f.read())

        encrypted = io.BytesIO()
        ms_file = msoffcrypto.OfficeFile(file_data)
        
        # Aplicar encriptación
        ms_file.encrypt(password, encrypted)

        # Sobrescribir el archivo original con los datos encriptados
        with open(file_path, "wb") as f:
            f.write(encrypted.getbuffer())
        
        print(f"🔒 Excel encriptado correctamente: {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"❌ Error al encriptar el Excel: {e}")

def protect_excel(file_path, password, user_password="1234", protect_sheet=True , blind=False):
    """
    Cifra un archivo Excel con una o dos capas de seguridad mediante el llamado
    a otras funciones definidas en este archivo.    
    Args:
        file_path (str): Ruta local del archivo Excel a proteger.
        password (str): Contraseña que se aplicará al archivo.
        protect_sheet (bool) default=True : Aplicar bloque de celdas al archivo xlsx.
        blind (bool) default=False : True para aplicar contraseña que nadie sabe (ciega).
                                    por defecto se aplica la contraseña del argumento password
    """
    if protect_sheet:
        if blind:
            # Protege el excel con contraseña desconocida para las celdas. 
            protect_sheet_only(file_path, blind=True)
            # Excel se abre con password del propietario(owner). 
            encryp_excel(file_path, password)
        else:
            # Excel solo modificable con contraseña de propietario(owner).
            protect_sheet_only(file_path, password)
            # Excel se abre con contraseña de usuario(User).
            encryp_excel(file_path, user_password)
    else:
        # Si no existe bloqueo de celdas.
        # El excel se abre con contraseña de propietario(owner).
        encryp_excel(file_path, password)
            
